"""Tests for strict parsing of organizer flow and signal workbooks."""

from datetime import time
from pathlib import Path

from openpyxl import Workbook

from traffic_platform.scenario_engine.official_models import MOVEMENTS
from traffic_platform.scenario_engine.official_workbook import parse_official_workbook


def _write_fixture(path: Path, *, incorrect_total: bool = False) -> None:
    workbook = Workbook()
    flow = workbook.active
    flow.title = "flow"
    sections = (
        (1, 4, 12, 13, 7, "早高峰时段流量数据"),
        (15, 18, 26, 27, 14, "平峰时段流量数据"),
        (29, 32, 40, 41, 17, "晚高峰时段流量数据"),
    )
    for section_row, start_row, total_row, ratio_row, hour, label in sections:
        flow.cell(section_row, 1, label)
        flow.cell(section_row + 1, 1, "统计开始时间")
        flow.cell(section_row + 1, 2, "统计结束时间")
        for column, approach in zip(
            (3, 6, 9, 12),
            ("东进口", "西进口", "南进口", "北进口"),
            strict=True,
        ):
            flow.cell(section_row + 1, column, approach)
            for offset, movement in enumerate(("左转(pcu)", "直行(pcu)", "右转(pcu)")):
                flow.cell(section_row + 2, column + offset, movement)
        for offset in range(8):
            row = start_row + offset
            minute = (offset * 15) % 60
            end_total_minutes = offset * 15 + 15
            flow.cell(row, 1, time(hour + (offset * 15) // 60, minute))
            flow.cell(
                row,
                2,
                time(hour + end_total_minutes // 60, end_total_minutes % 60),
            )
            for column in range(3, 15):
                flow.cell(row, column, 1)
        flow.cell(total_row, 1, "配置参数")
        flow.cell(total_row, 2, "总流量")
        flow.cell(ratio_row, 2, "转向比例")
        for column in (3, 6, 9, 12):
            flow.cell(total_row, column, 25 if incorrect_total else 24)
            for offset in range(3):
                ratio = flow.cell(ratio_row, column + offset, 1 / 3)
                ratio.number_format = "0.00_ "

    signal = workbook.create_sheet("signal")
    for row, label, window in (
        (3, "早高峰", "7:00-9:00"),
        (4, "平峰", "14:30-16:30"),
        (5, "晚高峰", "17:30-19:30"),
    ):
        signal.cell(row, 1, label)
        signal.cell(row, 2, window)
        signal.cell(row, 3, 1)
        signal.cell(row, 4, "测试相位")
        signal.cell(row, 5, 20)
        signal.cell(row, 6, 3)
        signal.cell(row, 7, 2)
        signal.cell(row, 8, 25)
        signal.cell(row, 9, 25)
    workbook.save(path)


def test_parser_preserves_exact_interval_counts_and_cycles(tmp_path: Path) -> None:
    workbook_path = tmp_path / "fixture.xlsx"
    _write_fixture(workbook_path)

    parsed = parse_official_workbook(workbook_path)

    assert set(parsed.demand_profiles) == {"am_peak", "offpeak", "pm_peak"}
    assert parsed.demand_profiles["am_peak"].total_vehicles == 8 * len(MOVEMENTS)
    assert parsed.demand_profiles["offpeak"].approach_totals == {
        "E": 24,
        "W": 24,
        "S": 24,
        "N": 24,
    }
    assert parsed.signal_profiles["pm_peak"].cycle_s == 25
    assert parsed.signal_profiles["pm_peak"].phases[0].total_s == 25


def test_parser_audits_inconsistent_source_totals_without_changing_demand(
    tmp_path: Path,
) -> None:
    workbook_path = tmp_path / "invalid.xlsx"
    _write_fixture(workbook_path, incorrect_total=True)

    parsed = parse_official_workbook(workbook_path)

    assert parsed.demand_profiles["am_peak"].total_vehicles == 8 * len(MOVEMENTS)
    assert parsed.source_audit["workbook_source_consistent"] is False
    assert parsed.source_audit["profiles"]["am_peak"]["flow"][
        "approach_total_differences"
    ]["E"] == {"source": 25.0, "calculated": 24}
