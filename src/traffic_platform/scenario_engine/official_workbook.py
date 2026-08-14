"""Strict parser for the organizer's 20-intersection Excel template."""

from datetime import time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from traffic_platform.scenario_engine.official_models import (
    MOVEMENTS,
    PROFILE_KEYS,
    DemandProfile,
    FlowInterval,
    OfficialWorkbook,
    SignalPhase,
    SignalProfile,
)

_APPROACH_LABELS = {
    "东进口": "E",
    "西进口": "W",
    "南进口": "S",
    "北进口": "N",
    "西南进口": "S",
    "东北进口": "N",
    "东南进口": "E",
    "西北进口": "W",
}
_PROFILE_LABELS = {
    "早高峰": "am_peak",
    "平峰": "offpeak",
    "晚高峰": "pm_peak",
}
_FLOW_SECTION_LABELS = {
    "早高峰时段流量数据": "am_peak",
    "平峰时段流量数据": "offpeak",
    "晚高峰时段流量数据": "pm_peak",
}


class OfficialWorkbookError(ValueError):
    """Raised when organizer data violates the declared workbook contract."""


def _as_nonnegative_int(value: Any, *, cell: str) -> int:
    if value is None:
        return 0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise OfficialWorkbookError(f"{cell} must be numeric, got {value!r}")
    rounded = round(float(value))
    if abs(float(value) - rounded) > 1e-9 or rounded < 0:
        raise OfficialWorkbookError(f"{cell} must be a non-negative integer PCU count")
    return int(rounded)


def _as_nonnegative_number(value: Any, *, cell: str) -> float:
    """Read a summary cell without forcing organizer formulas to be integral."""

    if value is None:
        return 0.0
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise OfficialWorkbookError(f"{cell} must be numeric, got {value!r}")
    number = float(value)
    if number < 0:
        raise OfficialWorkbookError(f"{cell} must be non-negative")
    return number


def _clock_text(value: Any, *, cell: str) -> str:
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, str) and value.strip():
        return value.strip()
    raise OfficialWorkbookError(f"{cell} must contain a clock time")


def _movement_columns(
    sheet: Worksheet,
    header_row: int,
) -> tuple[dict[str, int], tuple[str, ...]]:
    """Resolve directions from workbook headers instead of assuming N/S order."""

    movement_columns: dict[str, int] = {}
    approach_order: list[str] = []
    for start_column in (3, 6, 9, 12):
        raw_label = sheet.cell(header_row, start_column).value
        label = str(raw_label or "").strip()
        try:
            approach = _APPROACH_LABELS[label]
        except KeyError as exc:
            raise OfficialWorkbookError(
                f"unsupported approach header {label!r} at "
                f"{sheet.title}!{sheet.cell(header_row, start_column).coordinate}"
            ) from exc
        if approach in approach_order:
            raise OfficialWorkbookError(f"duplicate approach header {label!r}")
        approach_order.append(approach)
        for offset, turn in enumerate(("L", "S", "R")):
            movement_columns[f"{approach}_{turn}"] = start_column + offset
    if set(approach_order) != {"E", "W", "N", "S"}:
        raise OfficialWorkbookError(
            f"flow sheet must declare E/W/N/S once, got {approach_order}"
        )
    return movement_columns, tuple(approach_order)


def _flow_sections(sheet: Worksheet) -> dict[str, tuple[int, list[int], int, int]]:
    """Locate section rows because organizer workbooks do not share one layout."""

    section_starts: list[tuple[int, str]] = []
    for row in range(1, sheet.max_row + 1):
        label = str(sheet.cell(row, 1).value or "").strip()
        if label in _FLOW_SECTION_LABELS:
            section_starts.append((row, _FLOW_SECTION_LABELS[label]))
    if {key for _row, key in section_starts} != set(PROFILE_KEYS):
        raise OfficialWorkbookError(
            f"flow sheet must contain exactly three profile sections, got {section_starts}"
        )
    result: dict[str, tuple[int, list[int], int, int]] = {}
    for index, (section_row, key) in enumerate(section_starts):
        section_end = (
            section_starts[index + 1][0] - 1
            if index + 1 < len(section_starts)
            else sheet.max_row
        )
        header_row = next(
            (
                row
                for row in range(section_row + 1, min(section_end, section_row + 6) + 1)
                if str(sheet.cell(row, 1).value or "").strip() == "统计开始时间"
            ),
            None,
        )
        if header_row is None:
            raise OfficialWorkbookError(f"{key} has no flow header row")
        total_row = next(
            (
                row
                for row in range(header_row + 1, section_end + 1)
                if str(sheet.cell(row, 1).value or "").strip() == "配置参数"
            ),
            None,
        )
        if total_row is None:
            raise OfficialWorkbookError(f"{key} has no configuration total row")
        record_rows = [
            row
            for row in range(header_row + 2, total_row)
            if sheet.cell(row, 1).value is not None
            and sheet.cell(row, 2).value is not None
        ]
        if len(record_rows) != 8:
            raise OfficialWorkbookError(
                f"{key} must have eight 15-minute rows, got {record_rows}"
            )
        ratio_row = next(
            (
                row
                for row in range(total_row + 1, min(section_end, total_row + 3) + 1)
                if str(sheet.cell(row, 2).value or "").strip() == "转向比例"
            ),
            None,
        )
        if ratio_row is None:
            raise OfficialWorkbookError(f"{key} has no turn-ratio row")
        result[key] = (header_row, record_rows, total_row, ratio_row)
    return result


def _ratio_value(value: Any, *, cell: str) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise OfficialWorkbookError(f"{cell} must be numeric or blank, got {value!r}")
    ratio = float(value)
    if not 0 <= ratio <= 1:
        raise OfficialWorkbookError(f"{cell} must be between 0 and 1")
    return ratio


def _display_tolerance(number_format: str) -> float:
    """Return half of the last displayed decimal unit for an Excel format."""

    if "." not in number_format:
        return 1e-9
    fractional = number_format.split(".", maxsplit=1)[1]
    decimals = 0
    for character in fractional:
        if character not in {"0", "#"}:
            break
        decimals += 1
    return 0.5 * 10**-decimals + 1e-12 if decimals else 1e-9


def _ratios_differ(
    source: float | None,
    calculated: float | None,
    *,
    tolerance: float,
) -> bool:
    return (
        source is not None
        and calculated is not None
        and abs(source - calculated) > tolerance
    )


def _parse_demand(
    sheet: Worksheet,
) -> tuple[dict[str, DemandProfile], dict[str, Any]]:
    result: dict[str, DemandProfile] = {}
    audit: dict[str, Any] = {}
    labels = {"am_peak": "早高峰", "offpeak": "平峰", "pm_peak": "晚高峰"}
    for key, (header_row, record_rows, total_row, ratio_row) in _flow_sections(
        sheet
    ).items():
        movement_columns, approach_order = _movement_columns(sheet, header_row)
        intervals: list[FlowInterval] = []
        for interval_index, row in enumerate(record_rows):
            counts = {
                movement: _as_nonnegative_int(
                    sheet.cell(row, column).value,
                    cell=f"{sheet.title}!{sheet.cell(row, column).coordinate}",
                )
                for movement, column in movement_columns.items()
            }
            intervals.append(
                FlowInterval(
                    begin_s=interval_index * 900,
                    end_s=(interval_index + 1) * 900,
                    source_begin=_clock_text(
                        sheet.cell(row, 1).value,
                        cell=f"{sheet.title}!A{row}",
                    ),
                    source_end=_clock_text(
                        sheet.cell(row, 2).value,
                        cell=f"{sheet.title}!B{row}",
                    ),
                    counts=counts,
                )
            )
        movement_totals = {
            movement: sum(interval.counts[movement] for interval in intervals)
            for movement in MOVEMENTS
        }
        approach_totals = {
            approach: sum(
                movement_totals[f"{approach}_{turn}"] for turn in ("L", "S", "R")
            )
            for approach in ("E", "W", "S", "N")
        }
        source_approach_totals = {
            approach: _as_nonnegative_number(
                sheet.cell(total_row, column).value,
                cell=f"{sheet.title}!{sheet.cell(total_row, column).coordinate}",
            )
            for approach, column in {
                movement.split("_")[0]: column
                for movement, column in movement_columns.items()
                if movement.endswith("_L")
            }.items()
        }
        source_turn_ratios = {
            movement: _ratio_value(
                sheet.cell(ratio_row, column).value,
                cell=(
                    f"{sheet.title}!"
                    f"{sheet.cell(ratio_row, column).coordinate}"
                ),
            )
            for movement, column in movement_columns.items()
        }
        source_turn_tolerances = {
            movement: _display_tolerance(sheet.cell(ratio_row, column).number_format)
            for movement, column in movement_columns.items()
        }
        calculated_turn_ratios = {
            movement: (
                movement_totals[movement] / approach_totals[movement[0]]
                if approach_totals[movement[0]]
                else None
            )
            for movement in MOVEMENTS
        }
        ratio_differences = {
            movement: {
                "source": source_turn_ratios[movement],
                "calculated": calculated_turn_ratios[movement],
                "display_tolerance": source_turn_tolerances[movement],
            }
            for movement in MOVEMENTS
            if _ratios_differ(
                source_turn_ratios[movement],
                calculated_turn_ratios[movement],
                tolerance=source_turn_tolerances[movement],
            )
        }
        approach_differences = {
            approach: {
                "source": source_approach_totals[approach],
                "calculated": approach_totals[approach],
            }
            for approach in ("E", "W", "N", "S")
            if source_approach_totals[approach] != approach_totals[approach]
        }
        audit[key] = {
            "approach_header_order": list(approach_order),
            "source_rows": {
                "header": header_row,
                "records": record_rows,
                "totals": total_row,
                "ratios": ratio_row,
            },
            "approach_totals_match": not approach_differences,
            "approach_total_differences": approach_differences,
            "turn_ratios_match": not ratio_differences,
            "turn_ratio_differences": ratio_differences,
            "simulation_demand_authority": "eight_15_min_movement_rows",
        }
        result[key] = DemandProfile(
            key=key,
            label=labels[key],
            clock_window=f"{intervals[0].source_begin}-{intervals[-1].source_end}",
            intervals=tuple(intervals),
            movement_totals=movement_totals,
            approach_totals=approach_totals,
            source_approach_totals=source_approach_totals,
        )
    return result, audit


def _parse_signals(
    sheet: Worksheet,
) -> tuple[dict[str, SignalProfile], dict[str, Any]]:
    phases: dict[str, list[SignalPhase]] = {key: [] for key in PROFILE_KEYS}
    declared_phase_totals: dict[str, list[dict[str, Any]]] = {
        key: [] for key in PROFILE_KEYS
    }
    metadata: dict[str, tuple[str, str, int]] = {}
    current_key: str | None = None
    for row in range(3, sheet.max_row + 1):
        label_value = sheet.cell(row, 1).value
        if isinstance(label_value, str) and label_value.strip():
            label = label_value.strip()
            try:
                current_key = _PROFILE_LABELS[label]
            except KeyError as exc:
                raise OfficialWorkbookError(f"unsupported signal profile label {label!r}") from exc
            window = str(sheet.cell(row, 2).value or "").strip()
            cycle = _as_nonnegative_int(
                sheet.cell(row, 9).value,
                cell=f"{sheet.title}!I{row}",
            )
            if not window or cycle <= 0:
                raise OfficialWorkbookError(f"incomplete profile metadata at row {row}")
            metadata[current_key] = (label, window, cycle)
        phase_value = sheet.cell(row, 3).value
        if phase_value is None:
            continue
        if current_key is None:
            raise OfficialWorkbookError(f"phase at row {row} has no profile label")
        phase_id = str(_as_nonnegative_int(phase_value, cell=f"{sheet.title}!C{row}"))
        name = str(sheet.cell(row, 4).value or "").strip()
        if not name:
            raise OfficialWorkbookError(f"phase {phase_id} has no name")
        phase = SignalPhase(
            phase_id=phase_id,
            name=name,
            green_s=_as_nonnegative_int(sheet.cell(row, 5).value, cell=f"E{row}"),
            yellow_s=_as_nonnegative_int(sheet.cell(row, 6).value, cell=f"F{row}"),
            all_red_s=_as_nonnegative_int(sheet.cell(row, 7).value, cell=f"G{row}"),
        )
        stated_total = _as_nonnegative_int(sheet.cell(row, 8).value, cell=f"H{row}")
        declared_phase_totals[current_key].append(
            {
                "phase_id": phase_id,
                "source_total_s": stated_total,
                "calculated_total_s": phase.total_s,
                "matches": phase.total_s == stated_total,
            }
        )
        phases[current_key].append(phase)
    result: dict[str, SignalProfile] = {}
    audit: dict[str, Any] = {}
    for key in PROFILE_KEYS:
        if key not in metadata or not phases[key]:
            raise OfficialWorkbookError(f"signal profile {key} is missing")
        label, window, cycle = metadata[key]
        calculated_cycle = sum(phase.total_s for phase in phases[key])
        phase_total_mismatches = [
            item for item in declared_phase_totals[key] if not item["matches"]
        ]
        audit[key] = {
            "declared_cycle_s": cycle,
            "calculated_component_cycle_s": calculated_cycle,
            "cycle_matches": calculated_cycle == cycle,
            "phase_component_totals_match": not phase_total_mismatches,
            "phase_component_total_mismatches": phase_total_mismatches,
            "simulation_signal_authority": "phase_green_yellow_all_red_components",
        }
        result[key] = SignalProfile(
            key=key,
            label=label,
            clock_window=window,
            cycle_s=calculated_cycle,
            source_cycle_s=cycle,
            phases=tuple(phases[key]),
        )
    return result, audit


def parse_official_workbook(path: Path) -> OfficialWorkbook:
    """Parse and cross-check one organizer flow-and-signal workbook."""

    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if len(workbook.sheetnames) < 2:
            raise OfficialWorkbookError("workbook must contain flow and signal sheets")
        demand, demand_audit = _parse_demand(workbook[workbook.sheetnames[0]])
        signals, signal_audit = _parse_signals(workbook[workbook.sheetnames[1]])
    finally:
        workbook.close()
    profiles = {
        key: {
            "flow": demand_audit[key],
            "signal": signal_audit[key],
            "source_consistent": (
                demand_audit[key]["approach_totals_match"]
                and demand_audit[key]["turn_ratios_match"]
                and signal_audit[key]["cycle_matches"]
                and signal_audit[key]["phase_component_totals_match"]
            ),
        }
        for key in PROFILE_KEYS
    }
    source_audit = {
        "profiles": profiles,
        "workbook_source_consistent": all(
            profile["source_consistent"] for profile in profiles.values()
        ),
        "policy": (
            "simulation uses detailed 15-minute movement counts and per-phase "
            "green/yellow/all-red components; inconsistent summary cells are "
            "preserved as audit findings and never overwrite organizer files"
        ),
    }
    return OfficialWorkbook(
        path=path,
        demand_profiles=demand,
        signal_profiles=signals,
        source_audit=source_audit,
    )
